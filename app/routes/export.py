from fastapi import APIRouter, Depends, Query, HTTPException, Request
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.orm import Session
from datetime import datetime
import io
import csv
import typing as t

from app.database import get_db
from app import models

router = APIRouter(prefix="/api/export", tags=["export"])


def _require_premium_or_raise(request: Request, db: Session):
    # Prefer session plan; fall back to DB first user
    plan = None
    try:
        plan = request.session.get('user_plan')
    except Exception:
        plan = None
    if not plan:
        usuario = db.query(models.Usuario).first()
        if usuario:
            plan = usuario.plan
    if plan != 'premium':
        raise HTTPException(status_code=403, detail='Exportación en este formato disponible solo para usuarios Premium')


def _parse_date(s: str | None) -> datetime | None:
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            continue
    # try iso
    try:
        return datetime.fromisoformat(s)
    except Exception:
        return None


def _query_news(db: Session, start: datetime | None, end: datetime | None, source: str | None):
    q = db.query(models.Noticia)
    if start:
        q = q.filter(models.Noticia.created_at >= start)
    if end:
        q = q.filter(models.Noticia.created_at <= end)
    if source:
        like = f"%{source}%"
        q = q.filter(models.Noticia.fuente.ilike(like) | models.Noticia.url.ilike(like))
    return q.order_by(models.Noticia.created_at.desc()).all()


def _query_social(db: Session, start: datetime | None, end: datetime | None, source: str | None):
    q = db.query(models.SocialMediaPost)
    if start:
        q = q.filter(models.SocialMediaPost.created_at >= start)
    if end:
        q = q.filter(models.SocialMediaPost.created_at <= end)
    if source:
        like = f"%{source}%"
        q = q.filter(models.SocialMediaPost.platform.ilike(like) | models.SocialMediaPost.source.ilike(like))
    return q.order_by(models.SocialMediaPost.created_at.desc()).all()


@router.get("/json")
def export_json(
    request: Request,
    content: str = Query("all", description="all|news|social"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    source: str | None = Query(None),
    db: Session = Depends(get_db),
):
    start = _parse_date(start_date)
    end = _parse_date(end_date)
    if content in ("all", "news"):
        noticias = [
            {
                "id": n.id,
                "url": n.url,
                "fuente": n.fuente,
                "titulo": n.titulo,
                "categoria": n.categoria,
                "fecha_publicacion": n.fecha_publicacion.isoformat() if n.fecha_publicacion else None,
                "created_at": n.created_at.isoformat() if n.created_at else None,
                "updated_at": n.updated_at.isoformat() if n.updated_at else None,
            }
            for n in _query_news(db, start, end, source)
        ]
    else:
        noticias = []

    if content in ("all", "social"):
        social = [
            {
                "id": s.id,
                "platform": s.platform,
                "username": s.username,
                "text": s.text,
                "url": s.url,
                "likes": s.likes,
                "shares": s.shares,
                "created_at": s.created_at.isoformat() if s.created_at else None,
                "post_created_at": s.post_created_at.isoformat() if s.post_created_at else None,
                "source": s.source,
            }
            for s in _query_social(db, start, end, source)
        ]
    else:
        social = []

    # JSON export: premium only
    _require_premium_or_raise(request, db)
    return JSONResponse({"ok": True, "news": noticias, "social": social})


@router.get("/csv")
def export_csv(
    content: str = Query("all", description="all|news|social"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    source: str | None = Query(None),
    db: Session = Depends(get_db),
):
    start = _parse_date(start_date)
    end = _parse_date(end_date)

    def generate():
        buf = io.StringIO()
        writer = csv.writer(buf)

        if content in ("all", "news"):
            # header for news
            writer.writerow(["type", "id", "url", "fuente", "titulo", "categoria", "fecha_publicacion", "created_at", "updated_at", "snippet"])
            yield buf.getvalue()
            buf.seek(0)
            buf.truncate(0)
            for n in _query_news(db, start, end, source):
                snippet = (n.contenido or "")[:200].replace("\n", " ")
                writer.writerow(["news", n.id, n.url, n.fuente, n.titulo, n.categoria or "", n.fecha_publicacion.isoformat() if n.fecha_publicacion else "", n.created_at.isoformat() if n.created_at else "", n.updated_at.isoformat() if n.updated_at else "", snippet])
                yield buf.getvalue()
                buf.seek(0)
                buf.truncate(0)

        if content in ("all", "social"):
            # header for social
            writer.writerow(["type", "id", "platform", "username", "text", "url", "likes", "shares", "created_at", "post_created_at", "source"])
            yield buf.getvalue()
            buf.seek(0)
            buf.truncate(0)
            for s in _query_social(db, start, end, source):
                writer.writerow(["social", s.id, s.platform, s.username, (s.text or "").replace("\n", " "), s.url or "", s.likes or 0, s.shares or 0, s.created_at.isoformat() if s.created_at else "", s.post_created_at.isoformat() if s.post_created_at else "", s.source or ""])
                yield buf.getvalue()
                buf.seek(0)
                buf.truncate(0)

    filename = f"nexnews_export_{content}_{datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')}.csv"
    return StreamingResponse(generate(), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=\"{filename}\""})


@router.get("/xlsx")
def export_xlsx(
    request: Request,
    content: str = Query("all", description="all|news|social"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    source: str | None = Query(None),
    db: Session = Depends(get_db),
):
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception:
        raise HTTPException(status_code=500, detail="openpyxl is required to export XLSX. Install with 'pip install openpyxl'.")

    start = _parse_date(start_date)
    end = _parse_date(end_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "export"

    row = 1
    if content in ("all", "news"):
        headers = ["id", "url", "fuente", "titulo", "categoria", "fecha_publicacion", "created_at", "updated_at", "snippet"]
        ws.append(["type"] + headers)
        for n in _query_news(db, start, end, source):
            snippet = (n.contenido or "")[:500]
            ws.append(["news", n.id, n.url, n.fuente, n.titulo, n.categoria or "", n.fecha_publicacion.isoformat() if n.fecha_publicacion else "", n.created_at.isoformat() if n.created_at else "", n.updated_at.isoformat() if n.updated_at else "", snippet])

    if content in ("all", "social"):
        headers = ["id", "platform", "username", "text", "url", "likes", "shares", "created_at", "post_created_at", "source"]
        ws.append(["type"] + headers)
        for s in _query_social(db, start, end, source):
            ws.append(["social", s.id, s.platform, s.username, s.text or "", s.url or "", s.likes or 0, s.shares or 0, s.created_at.isoformat() if s.created_at else "", s.post_created_at.isoformat() if s.post_created_at else "", s.source or ""])

    # Auto-size columns (simple heuristic)
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
            except Exception:
                val = ""
            if len(val) > max_length:
                max_length = len(val)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 100)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    # XLSX export: premium only
    _require_premium_or_raise(request, db)
    filename = f"nexnews_export_{content}_{datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')}.xlsx"
    return StreamingResponse(bio, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": f"attachment; filename=\"{filename}\""})
import csv, io, json
import pandas as pd
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy import select
from sqlalchemy.orm import Session
from typing import Optional
from app.database import get_db
from app.models import Noticia

router = APIRouter()

def base_query(db: Session, q: Optional[str], fuente: Optional[str]):
    stmt = select(Noticia).order_by(Noticia.created_at.desc())
    if fuente:
        stmt = stmt.where(Noticia.fuente == fuente)
    if q:
        like = f"%{q}%"
        stmt = stmt.where((Noticia.titulo.ilike(like)) | (Noticia.contenido.ilike(like)))
    return db.scalars(stmt)

@router.get("/export/csv")
def export_csv(
    q: Optional[str] = Query(None),
    fuente: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    rows = base_query(db, q, fuente).all()
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["id", "url", "fuente", "titulo", "contenido", "fecha_publicacion", "imagen_path", "created_at", "updated_at"])
    for r in rows:
        writer.writerow([r.id, r.url, r.fuente, r.titulo, r.contenido, r.fecha_publicacion, r.imagen_path, r.created_at, r.updated_at])
    buf.seek(0)
    return StreamingResponse(iter([buf.read()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=news.csv"})

@router.get("/export/json")
def export_json(
    request: Request,
    q: Optional[str] = Query(None),
    fuente: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    # JSON export: premium only
    _require_premium_or_raise(request, db)
    rows = base_query(db, q, fuente).all()
    payload = [
        {
            "id": r.id,
            "url": r.url,
            "fuente": r.fuente,
            "titulo": r.titulo,
            "contenido": r.contenido,
            "fecha_publicacion": r.fecha_publicacion.isoformat() if r.fecha_publicacion else None,
            "imagen_path": r.imagen_path,
            "created_at": r.created_at.isoformat(),
            "updated_at": r.updated_at.isoformat(),
        }
        for r in rows
    ]
    return JSONResponse(payload)


@router.get("/export/xlsx")
def export_xlsx(
    request: Request,
    q: Optional[str] = Query(None),
    fuente: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Exportar noticias a Excel (.xlsx) usando pandas"""
    # XLSX export: premium only
    _require_premium_or_raise(request, db)
    rows = base_query(db, q, fuente).all()
    data = []
    for r in rows:
        data.append({
            "id": r.id,
            "url": r.url,
            "fuente": r.fuente,
            "titulo": r.titulo,
            "contenido": r.contenido,
            "fecha_publicacion": r.fecha_publicacion.isoformat() if r.fecha_publicacion else None,
            "imagen_path": r.imagen_path,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "updated_at": r.updated_at.isoformat() if r.updated_at else None,
        })

    df = pd.DataFrame(data)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Noticias')
    buf.seek(0)
    headers = {"Content-Disposition": "attachment; filename=news.xlsx"}
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
